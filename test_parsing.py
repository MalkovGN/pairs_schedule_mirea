import time
import os
import requests
import pyexcel
from bs4 import BeautifulSoup as Soup
from openpyxl import load_workbook
# import win32com.client as win32

from schedule.schedule_app.utils import xls_as_xlsx


def get_links():
    """
    Getting the link of the schedule to download it
    """
    url = 'https://www.mirea.ru/schedule/'
    page = requests.get(url)
    soup = Soup(page.text, "html.parser")
    schedule_link = soup.find_all("a", class_='uk-link-toggle')

    links_lst = []
    for elem in schedule_link:
        links_lst.append(elem.get('href'))
    return links_lst  # Return the list of the links of the Excel files


def get_files():
    """
    Download files for the next parsing
    Creating a list of the file names
    """
    file_names = []
    links_lst = get_links()
    for link in links_lst:
        url_to_download = link
        # url_to_download = links_lst[0]  # For test take the first link from the lst
        filename = url_to_download.split('/')[-1]  # Need to append a loop for checking all the files
        file_names.append(filename)

        file = requests.get(url_to_download, allow_redirects=True)
        open(f'{os.getcwd()}/schedule/schedule_app/static/schedule_app/{filename}', 'wb').write(file.content)

    xls_as_xlsx()

    directory = f'{os.getcwd()}/schedule/schedule_app/static/schedule_app/'
    for f in os.listdir(directory):
        os.remove(os.path.join(directory, f))

    # print(file_names)
    return file_names  # Return the list of the filenames after parsing


def read_file_info():
    # file_names = get_files()
    group_schedule = {}
    for file in os.listdir(os.getcwd() + '/excel_files/'):
        file_name, fileextantion = os.path.splitext(file)
        print(file_name, fileextantion)
        wb = load_workbook(f'{os.getcwd()}/excel_files/{file_name + fileextantion}')
        sheets_name_list = wb.sheetnames
    # group_names = []

        for sheet_name in sheets_name_list:
            sheet = wb[f'{sheet_name}']

            for column in range(6, sheet.max_column):
                group_name = sheet.cell(row=2, column=column).value
                CHOISES = ('Предмет', 'Дисциплина',)
                subjects_names = []
                if sheet.cell(row=3, column=column).value in CHOISES:
                    for row in range(4, 75+1):
                        subject_name = sheet.cell(row=row, column=column).value
                        if subject_name is not None:
                            subject_name = subject_name.replace('\n', '')
                        subjects_names.append(subject_name)
                    # group_schedule[group_name] = ''
                if group_name != '':
                    group_schedule[group_name] = subjects_names

    # print(group_names)
    print(group_schedule)


# #         # return group_schedule


def parse_big_list():
    # file = os.getcwd() + '/excel_files/KBiSP_1_kurs_1sem_magistry_22.09.2022.xlsx/'
    wb = load_workbook(f'{os.getcwd()}/excel_files/KBiSP_1_kurs_1sem_magistry_22.09.2022.xlsx')
    sheet_name_list = wb.sheetnames
    group_schedule = {}

    for sheet_name in sheet_name_list:
        sheet = wb[f'{sheet_name}']
        n = 4
        for column in range(1, sheet.max_column):
            CHOISES = ('Предмет', 'Дисциплина',)
            if sheet.cell(row=2, column=column).value == 'Группа':

                group_name = sheet.cell(row=2, column=column + n).value
                print(group_name)
                if sheet.cell(row=3, column=column).value in CHOISES:
                    subject_names = []
                    for row in range(4, 75 + 1):
                        subject_name = sheet.cell(row=row, column=column).value
                        if subject_name is None:
                            continue
                        else:
                            subject_name = subject_name.replace('\n', '')
                        subject_names.append(subject_name)
                    group_schedule[group_name] = subject_names
                n += 5
                # group_names.append(sheet.cell(row=2, column=column + 4).value)
                # group_names.append(sheet.cell(row=2, column=column + 9).value)
                # group_names.append(sheet.cell(row=2, column=column + 14).value)


def parse_pairs():
    # file = os.getcwd() + '/excel_files/KBiSP_1_kurs_1sem_magistry_22.09.2022.xlsx/'
    wb = load_workbook(f'{os.getcwd()}/excel_files/KBiSP_1_kurs_1sem_magistry_22.09.2022.xlsx')
    sheet_name_list = wb.sheetnames
    group_schedule = {}
    CHOISES = ('Предмет', 'Дисциплина',)

    for sheet_name in sheet_name_list:
        sheet = wb[f'{sheet_name}']
        for column in range(1, sheet.max_column):
            if sheet.cell(row=3, column=column).value in CHOISES:
                subjects_names = []
                for row in range(4, 87 + 1):
                    subject_name = sheet.cell(row=row, column=column).value
                    print(subject_name)
                    if subject_name is None:
                        continue
                    else:
                        subject_name = subject_name.replace('\n', '')
                        subjects_names.append(subject_name)
                print(subjects_names)
if __name__ == '__main__':
    # get_links()
    # get_files()
    # start_time = time.time()
    # read_file_info()
    # print(time.time() - start_time)

    # parse_big_list()
    parse_pairs()