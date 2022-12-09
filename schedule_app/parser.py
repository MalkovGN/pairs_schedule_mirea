from openpyxl import load_workbook


def check_empty_list(sheet):
    return True if sheet.cell(row=4, column=5).value is None else False


def rows_in_table(sheet):

    for row in range(4, sheet.max_row):
        if sheet.cell(row=row, column=5).value == 'II' and sheet.cell(row=row + 1, column=5).value is None:
            return row


def get_group_names(sheet):
    group_names = {}

    for column in range(2, sheet.max_column):
        if sheet.cell(row=2, column=column).value == 'Группа':
            if sheet.cell(row=2, column=column + 4).value is not None:
                group_names[sheet.cell(row=2, column=column + 4).value[0:10]] = column + 4
            if sheet.cell(row=2, column=column + 9).value is not None:
                group_names[sheet.cell(row=2, column=column + 9).value[0:10]] = column + 9

    return group_names


def get_days_length(sheet):
    rows_in_schedule_table = rows_in_table(sheet)

    days_row_dict = {}

    for row in range(4, rows_in_schedule_table + 1):
        days_row_dict[sheet.cell(row=row, column=1).value] = row

    days_row_dict = dict(sorted(days_row_dict.items(), key=lambda item: item[1]))
    rows = list(days_row_dict.keys())
    try:
        for index, key in enumerate(days_row_dict):
            if key == 'СУББОТА':
                days_row_dict[key] = [days_row_dict[rows[index]], rows_in_schedule_table]
            else:
                days_row_dict[key] = [days_row_dict[rows[index]], days_row_dict[rows[index + 1]] - 1]
    except IndexError:
        days_row_dict.popitem()

    return days_row_dict


def get_schedule(filepath):
    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    file_schedule = {}

    for sheet_name in sheet_names:
        sheet = wb[f'{sheet_name}']

        if check_empty_list(sheet):
            break

        group_names = get_group_names(sheet)
        days_row = get_days_length(sheet)

        group_schedule = {}

        for group in group_names.keys():
            week_schedule = {}
            for day in days_row.keys():
                subject_names = []
                for row in range(days_row[day][0], days_row[day][1] + 1):
                    pair_info = []
                    subject_name = sheet.cell(row=row, column=group_names[group]).value

                    if subject_name is None or subject_name == '':
                        continue
                    elif subject_name is not None:
                        subject_name = subject_name.replace('\n', ' ')
                        week_parity = sheet.cell(row=row, column=5).value
                        if week_parity == 'II':
                            time_start = sheet.cell(row=row - 1, column=3).value
                            time_end = sheet.cell(row=row - 1, column=4).value
                            lesson_number = sheet.cell(row=row - 1, column=2).value
                        else:
                            time_start = sheet.cell(row=row, column=3).value
                            time_end = sheet.cell(row=row, column=4).value
                            lesson_number = sheet.cell(row=row, column=2).value
                        lesson_type = sheet.cell(row=row, column=group_names[group] + 1).value
                        teacher = sheet.cell(row=row, column=group_names[group] + 2).value
                        cabinet = sheet.cell(row=row, column=group_names[group] + 3).value

                        pair_info.append(subject_name)
                        pair_info.append(time_start + ':' + time_end)
                        pair_info.append(lesson_number)
                        pair_info.append(lesson_type)
                        pair_info.append(teacher)
                        pair_info.append(cabinet)
                        pair_info.append(week_parity)

                    subject_names.append(pair_info)
                week_schedule[day] = subject_names
            group_schedule[group] = week_schedule
        file_schedule[sheet_name] = group_schedule

    return file_schedule
