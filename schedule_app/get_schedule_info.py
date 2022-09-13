import os
import requests
from bs4 import BeautifulSoup as Soup
from openpyxl import load_workbook


def get_links():
    """
    Getting the link of the schedule to download it
    """
    url = 'https://www.mirea.ru/schedule/'
    page = requests.get(url)
    soup = Soup(page.text, "html.parser")
    schedule_link = soup.find_all("a", class_='uk-link-toggle')

    links_lst = []
    for elem in schedule_link:
        links_lst.append(elem.get('href'))

    return links_lst  # Return the list of the links of the Excel files


def get_files():
    """
    Download files for the next parsing
    Creating a list of the file names
    """
    file_names = []
    links_lst = get_links()
    url_to_download = links_lst[0]  # For test take the first link from the lst
    filename = url_to_download.split('/')[-1]  # Need to append a loop for checking all the files
    file_names.append(filename)

    file = requests.get(url_to_download, allow_redirects=True)
    open(f'{os.getcwd()}/schedule_app/static/schedule_app/{filename}', 'wb').write(file.content)

    return file_names  # Return the list of the filenames after parsing


def read_file_info():
    file_names = get_files()
    wb = load_workbook(f'schedule_app/static/schedule_app/{file_names[0]}')
    sheets_name_list = wb.sheetnames
    group_names = []
    for sheet_name in sheets_name_list:
        sheet = wb[f'{sheet_name}']

        for column in range(6, sheet.max_column, 5):
            group_names.append(sheet.cell(row=2, column=column).value)

    return group_names  # Getting a group names from the one of the all sites 
