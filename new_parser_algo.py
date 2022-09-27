import time
import os
import requests
import pyexcel
from bs4 import BeautifulSoup as Soup
from openpyxl import load_workbook
# import win32com.client as win32

from schedule.schedule_app.utils import xls_as_xlsx


def rows_in_table(filepath):
    """
    Вычисляет количество строк, занимаемых
    таблицей с расписанием
    :param filepath:
    :return: словарь с названием листа и количеством строк
    """
    sheet_names_max_row = {}
    wb = load_workbook(filepath)
    sheet_name_list = wb.sheetnames

    for sheet_name in sheet_name_list:
        sheet = wb[f'{sheet_name}']
        for row in range(4, sheet.max_row):
            if sheet.cell(row=row, column=5).value == 'II' and sheet.cell(row=row + 1, column=5).value is None:
                sheet_names_max_row[sheet_name] = row

    return sheet_names_max_row


def parse_schedule():
    """
    Для примера берем один файл,
    далее добавить цикл.
    Открываем файл, с которым будем работать.
    """

    time_dict = {
        1: '9:00-10:30',
        2: '10:40-12:10',
        3: '12:40-14:10',
        4: '14:20-15:50',
        5: '16:20-17:50',
        6: '18:00-19:30',
        7: '19:40-21:10',
        8: '21:20-22:50',
    }

    WEEKDAYS = (
        'ПОНЕДЕЛЬНИК',
        'ВТОРНИК',
        'СРЕДА',
        'ЧЕТВЕРГ',
        'ПЯТНИЦА',
        'СУББОТА',
    )

    filepath = f'{os.getcwd()}/excel_files/ITU_2_k_Stromynka_22_23.xlsx'
    wb = load_workbook(filepath)
    sheet_name_list = wb.sheetnames
    group_schedule = {}

    for sheet_name in sheet_name_list: # Цикл по всем листам в файле
        sheet = wb[f'{sheet_name}']
        group_names = []

        """Добавить проверку на то, сколько групп на одном листе"""

        for column in range(1, sheet.max_column):
            if sheet.cell(row=2, column=column).value == 'Группа':
                if sheet.cell(row=2, column=column + 4).value is not None:
                    group_names.append([sheet.cell(row=2, column=column + 4).value[0:10], column + 4])
                if sheet.cell(row=2, column=column + 9).value is not None:
                    group_names.append([sheet.cell(row=2, column=column + 9).value[0:10], column + 9])
                # group_names.append([sheet.cell(row=2, column=column + 14).value[0:10], column + 14])

        rows_table = rows_in_table(filepath)[sheet_name]

        for group in group_names:
            subject_names = []

            for row in range(4, rows_table + 1):
                day_info = []
                daily_dict = {}
                # if sheet.cell(row=row, column=1).value in WEEKDAYS:
                #     # print(sheet.cell(row=row, column=1).value in WEEKDAYS)
                #     # print(sheet.cell(row=row, column=1).value)
                #     current_day = sheet.cell(row=row, column=1).value
                #     daily_dict[current_day] = None
                #     # print(daily_dict)
                subject_name = sheet.cell(row=row, column=group[1]).value
                if subject_name is None:
                    continue
                else:
                    # if sheet.cell(row=row, column=1).value in WEEKDAYS:

                    day_info.append(subject_name)
                    subject_type = sheet.cell(row=row, column=group[1] + 1).value
                    day_info.append(subject_type)
                    teacher_name = sheet.cell(row=row, column=group[1] + 2).value
                    day_info.append(teacher_name)
                    cabinet = sheet.cell(row=row, column=group[1] + 3).value
                    day_info.append(cabinet)
                    week_parity = sheet.cell(row=row, column=5).value
                    day_info.append(week_parity)
                    if week_parity == 'II':
                        lesson_number = sheet.cell(row=row-1, column=2).value
                        day_info.append(lesson_number)
                    else:
                        lesson_number = sheet.cell(row=row, column=2).value
                        day_info.append(lesson_number)
                    lesson_time = time_dict[int(lesson_number)]
                    day_info.append(lesson_time)
                    # if int(lesson_number) == 1:
                    #     daily_dict[sheet.cell(row=row, column=1).value] = None
                    # daily_dict[sheet.cell(row=row, column=1).value] = day_info
                    subject_names.append(day_info)
                    # print(list(daily_dict.keys()))
                    # daily_dict[current_day] = day_info

            group_schedule[group[0]] = subject_names

    print(group_schedule)

    print(rows_in_table(filepath))
    # print(group_names)
    # print(sheet.max_row)



    """
    Определение количества строк в таблице с расписанием
    Тк внизу файла лежит еще всякий шлак 
    """


if __name__ == '__main__':
    parse_schedule()